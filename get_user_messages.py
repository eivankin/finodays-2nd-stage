from openpyxl import load_workbook
# from openpyxl.cell import Cell

if __name__ == '__main__':
    wb = load_workbook('data/FINODAYS_Доп. материал для Почта Банк_Диалоги.xlsx')
    for sn in wb.sheetnames:
        print(sn)
        marks = []
        for row in wb[sn]:
            if row[1].value == 'CLIENT':
                print(row[2].value)
            elif row[2].value.startswith('Оценка'):
                marks.append(row[2].value)
        print('Оценки:', *marks, end='\n' + '-' * 50 + '\n')
